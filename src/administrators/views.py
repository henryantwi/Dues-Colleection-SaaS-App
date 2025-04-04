import csv

from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.cache import never_cache
from icecream import ic
from .sms_client import send_sms_get
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


from payments.models import Payment
from students.models import Student, Department

from .forms import EmailAuthenticationForm

ic.disable()

def login_view(request):
    if request.method == "POST":
        form = EmailAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f"Welcome back, {user.username.capitalize()}! 👋")
            return redirect("home")
        else:
            messages.error(request, "Invalid email or password.")
    else:
        form = EmailAuthenticationForm()
    return render(request, "administrators/login.html", {"form": form})


def logout_view(request):
    name: str = request.user.username.capitalize()
    logout(request)
    # messages.info(request, "You have been logged out.")
    messages.info(request, f"Good bye, {name}👋.")
    return redirect("administrators:login")


@login_required
def student_search(request):
    query = request.GET.get("q", "")
    if query:
        students = Student.objects.filter(
            Q(ref_number__icontains=query)
            | Q(full_name__icontains=query)
            | Q(email__icontains=query)
            | Q(payment__reference__icontains=query)
            | Q(unique_code__icontains=query)
        ).select_related("department", "payment")

        # check if students are in the same department as the admin
        if not request.user.is_superuser:
            students = students.filter(department=request.user.department)
    else:
        students = Student.objects.none()

    return render(
        request, "administrators/search.html", {"students": students, "query": query}
    )


def get_time_ago(timestamp):
    """Convert timestamp to '2m ago' format"""
    now = timezone.now()
    diff = now - timestamp

    seconds = diff.total_seconds()
    if seconds < 60:
        return "just now"
    elif seconds < 3600:
        minutes = int(seconds / 60)
        return f"{minutes}m ago"
    elif seconds < 86400:
        hours = int(seconds / 3600)
        return f"{hours}h ago"
    else:
        days = int(seconds / 86400)
        return f"{days}d ago"


@login_required
def admin_dashboard(request):
    # Check if user is a department admin or superuser
    department = None

    if hasattr(request.user, "departmentadmin"):
        department_admin = request.user.departmentadmin
        if not department_admin.is_active:
            raise PermissionDenied
        department = department_admin.department
    elif not request.user.is_superuser:
        raise PermissionDenied

        ic(department)
    # Get department statistics
    if department:
        # For department admins, only show their department's data
        students_query = Student.objects.filter(department=department)
        successful_payments = Payment.objects.filter(
            department=department,
            status="Successful",
            students__isnull=False,  # Only count payments with associated students
        ).distinct()

        # Calculate statistics for the specific department
        total_amount = successful_payments.aggregate(total=Sum("amount"))["total"] or 0
        total_students = students_query.count()
        paid_students = (
            students_query.filter(payment__status="Successful").distinct().count()
        )
        pending_payments = (
            Payment.objects.filter(
                department=department, status="Pending", students__isnull=False
            )
            .distinct()
            .count()
        )

        # Get recent payments for this department only
        recent_payments = (
            Payment.objects.filter(
                department=department,
                students__department=department,  # Ensure students match department
            )
            .exclude(students=None)  # Only show payments with students
            .select_related("department")
            .prefetch_related("students")
            .order_by("-created_at")
            .distinct()[:5]
        )
    else:
        # For superusers, show data from all departments
        students_query = Student.objects.all()
        successful_payments = Payment.objects.filter(
            status="Successful", students__isnull=False
        ).distinct()

        # Calculate overall statistics
        total_amount = successful_payments.aggregate(total=Sum("amount"))["total"] or 0
        total_students = students_query.count()
        paid_students = (
            Student.objects.filter(payment__status="Successful").distinct().count()
        )
        pending_payments = (
            Payment.objects.filter(status="Pending", students__isnull=False)
            .distinct()
            .count()
        )

        # Get recent payments from all departments
        recent_payments = (
            Payment.objects.exclude(students=None)
            .select_related("department")
            .prefetch_related("students")
            .order_by("-created_at")[:5]
        )

    # Add department statistics for superadmin
    department_stats = None
    if request.user.is_superuser:
        department_stats = []
        departments = Department.objects.filter(is_active=True)
        for dept in departments:
            dept_students = Student.objects.filter(department=dept)
            dept_stats = {
                'name': dept.full_name or dept.name,
                'total_students': dept_students.count(),
                'paid_students': dept_students.filter(payment__status="Successful").distinct().count(),
            }
            department_stats.append(dept_stats)

    # Format recent payments for display with additional checks
    recent_activities = []
    for payment in recent_payments:
        student = payment.get_student()
        # Additional check to ensure student belongs to admin's department
        if student and (not department or student.department == department):
            recent_activities.append(
                {
                    "type": "payment",
                    "status": payment.status,
                    "amount": payment.amount,
                    "student_name": student.full_name,
                    "student_ref": student.ref_number,
                    "time_ago": get_time_ago(payment.created_at),
                    "created_by": "System",
                    "department": payment.department.name,
                }
            )

    context = {
        "department": department,
        "total_amount": total_amount,
        "total_students": total_students,
        "paid_students": paid_students,
        "pending_payments": pending_payments,
        "recent_activities": recent_activities,
        "department_stats": department_stats,  # Add to context
    }

    return render(request, "administrators/admin_dashboard.html", context)


@never_cache
@login_required
def student_detail(request, ref_number):
    student = get_object_or_404(Student, ref_number=ref_number)

    # Check if user/admin has permIission to view this student
    if not request.user.is_superuser and request.user.department != student.department:
        raise Http404("Student not found")

    # Get all payments for this student
    payments = Payment.objects.filter(students=student).order_by("-created_at")

    context = {
        "student": student,
        "payments": payments,
        "latest_payment": payments.first() if payments.exists() else None,
    }

    return render(request, "administrators/student_detail.html", context)


@login_required
def mark_payment_successful(request, payment_ref):
    payment = get_object_or_404(Payment, reference=payment_ref)
    student = get_object_or_404(Student, payment=payment)

    # Check permissions
    if not request.user.is_superuser:
        try:
            department_admin = request.user.departmentadmin
            if (
                not department_admin.is_active
                or department_admin.department != payment.department
            ):
                raise PermissionDenied
        except:
            raise PermissionDenied

    if payment.status == "Pending":
        payment.status = "Successful"
        payment.save()
        number = f"233{student.mobile[1:]}"
        try:
            first_name = student.full_name.split()[0]
            message = (
            f"Dear {first_name}, your payment (Ref: {payment.reference}) "
            f"has been successfully confirmed. Thank you for your payment."
            )
        except (AttributeError, IndexError):
            message = (
            f"Dear Student, your payment (Ref: {payment.reference}) "
            f"has been successfully confirmed. Thank you for your payment."
            )
        if payment.method == "Cash":
            message += " (Cash Payment)"
        send_sms_get([number], message)
        messages.success(request, "Payment has been marked as successful.")
    else:
        messages.error(request, "This payment cannot be modified.")

    return redirect(
        "administrators:student_detail", ref_number=payment.students.first().ref_number
    )


@login_required
def student_list(request):
    query = request.GET.get("q", "")

    if request.user.is_superuser:
        students = Student.objects.select_related("department", "payment").all()
    else:
        department = request.user.departmentadmin.department
        students = Student.objects.select_related("department", "payment").filter(
            department=department
        )

    # Apply search filter if query exists
    if query:
        students = students.filter(
            Q(ref_number__icontains=query)
            | Q(full_name__icontains=query)
            | Q(email__icontains=query)
            | Q(payment__reference__icontains=query)
            | Q(unique_code__icontains=query)
        )

    students = students.order_by("-created_at")

    # Pagination - 10 per page
    paginator = Paginator(students, 10)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "administrators/student_list.html",
        {"page_obj": page_obj, "query": query},
    )


@login_required
def download_students_excel(request):
    # Create a workbook and active sheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"

    # Query students based on the user type
    if request.user.is_superuser:
        students = Student.objects.select_related("department", "payment").all()
    else:
        department = request.user.departmentadmin.department
        students = Student.objects.select_related("department", "payment").filter(
            department=department
        )

    # Base header
    header = [
        "Reference Number",
        "Full Name",
        "Email",
        "Mobile",
        "Department",
        "Level",
        "Payment Status",
        "Payment Reference",
        "Amount Paid",
        "Registration Date",
    ]

    # Check if T-shirt status should be included
    include_tshirt_status = any(student.department.tshirt_included for student in students)

    # Add "T-shirt Status" to the header if needed
    if include_tshirt_status:
        header.append("T-shirt Status")

    # Write the header row
    for col_num, column_title in enumerate(header, 1):
        sheet.cell(row=1, column=col_num, value=column_title)

    # Write student data rows
    for row_num, student in enumerate(students, start=2):
        payment_status = "No Payment"
        payment_reference = "N/A"
        payment_amount = "0.00"

        # Payment details if available
        if hasattr(student, "payment") and student.payment:
            payment_status = student.payment.status
            payment_reference = student.payment.reference
            payment_amount = student.payment.amount
            # Subtract service charge for cash payments
            if student.payment.method == "Cash":
                payment_amount = payment_amount - student.department.service_charge

        # Base row data
        row_data = [
            student.ref_number,
            student.full_name,
            student.email,
            student.mobile,
            student.department.name,
            student.level,
            payment_status,
            payment_reference,
            payment_amount,
            student.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        ]

        # Add T-shirt status if applicable
        if include_tshirt_status:
            tshirt_status = "N/A"
            if student.department.tshirt_included:
                if student.year_group == 1:
                    if student.tshirt_option == "full":
                        tshirt_status = "Full Payment"
                    elif student.tshirt_option == "partial":
                        tshirt_status = "Partial Payment"
            row_data.append(tshirt_status)

        # Write data to the Excel sheet
        for col_num, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    # Adjust column widths for better readability
    for col_num, column_title in enumerate(header, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = max(len(str(column_title)), 15)

    # Prepare HTTP response with Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="students.xlsx"'
    workbook.save(response)

    return response


@login_required
def update_tshirt_payment(request, ref_number):
    student = get_object_or_404(Student, ref_number=ref_number)
    
    # Check if user has permission
    if not request.user.is_superuser and request.user.department != student.department:
        raise PermissionDenied
    
    # Verify this is a valid update case
    if student.year_group == 1 and student.department.tshirt_included and student.tshirt_option == 'partial':
        student.tshirt_option = 'full'
        student.save()
        number = f"233{student.mobile[1:]}"
        try:
            first_name = student.full_name.split()[0]
            message = (
            f"Dear {first_name}, your T-shirt payment has been successfully "
            "updated to full payment. You can now collect your T-shirt. Thank you!"
            )
        except (AttributeError, IndexError):
            message = (
            "Dear Student, your T-shirt payment has been successfully "
            "updated to full payment. You can now collect your T-shirt. Thank you!"
            )
        try:
            send_sms_get([number], message)
        except Exception as e:
            messages.warning(request, f"Payment updated but SMS notification failed: {str(e)}")
        messages.success(request, "T-shirt payment has been updated to full payment.")
    else:
        messages.error(request, "Invalid T-shirt payment update request.")
    
    return redirect('administrators:student_detail', ref_number=student.ref_number)


@login_required
def filtered_student_list(request, filter_type):
    if request.user.is_superuser:
        students = Student.objects.select_related("department", "payment").all()
    else:
        department = request.user.departmentadmin.department
        students = Student.objects.select_related("department", "payment").filter(
            department=department
        )

    # Apply filters based on filter_type
    if filter_type == 'paid':
        students = students.filter(payment__status='Successful')
    elif filter_type == 'pending':
        students = students.filter(payment__status='Pending', payment__method='Cash')
    # 'all' doesn't need additional filtering

    # Sort by creation date
    students = students.order_by('-created_at')

    # Pagination - 10 per page
    paginator = Paginator(students, 10)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "filter_type": filter_type,
        "query": "",  # Empty query for search form
    }
    
    return render(request, "administrators/student_list.html", context)
